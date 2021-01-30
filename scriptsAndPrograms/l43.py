import openpyxl
wb = openpyxl.Workbook()
#wb
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'].value
sheet['A1'].value == None

sheet['A1'] = 42
sheet['A1'] = "Hello"

import os 
os.chdir('..\\textFile')
wb.save('example1.xlsx')
sheet2 = wb.create_sheet()
wb.get_sheet_names()

sheet2.textFile = "My New Sheet Name"
wb.save('example2.xlsx')
wb.create_sheet(index = 0, title='My Other Sheet')
wb.save('example3.xlsx')