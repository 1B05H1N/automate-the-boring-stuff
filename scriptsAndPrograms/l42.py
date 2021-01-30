import openpyxl
import os

os.chdir('..\\textFile')

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')
type(sheet)

workbook.get_sheet_by_names()

sheet['A1']
cell = sheet['A1']
cell.value

str(cell.value)
str(sheet['A1'].value)
str(sheet['B1'].value)
str(sheet['C1'].value)

sheet.cell(row=1, column=2)

sheet['B1']

for i in range(1,8):
	print(i, sheet.cell(row=i, column=2).value)

	